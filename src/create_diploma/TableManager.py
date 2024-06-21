import asyncio
import os
import openpyxl  # .xlsx,.xlsm,.xltx,.xltm - поддерживаемые форматы
from PIL import ImageDraw, Image, ImageFont
from openpyxl.workbook import Workbook
from openpyxl.worksheet.worksheet import Worksheet

from api.src.const import BASE_PATH
from api.src.create_diploma.ElementFont import find_ttf_file
from api.src.create_diploma.download_result import get_result
from api.src.modules import ElementsList, ElementWrapper, Element

from fastapi import HTTPException

def insert_text_on_image():
    pass


def find_key_index(sheet: Worksheet, elements: ElementsList) -> dict:
    """"""
    keys_index = dict()
    element: ElementWrapper
    for element in elements.elements:
        for row in sheet.iter_rows(values_only=True):
            index = 0
            for value in row:
                if value == element.key:
                    keys_index[element.key] = index
                index += 1

    return keys_index


async def print_excel_rows(user_name: str, project_name: str, elements: ElementsList, table_format: str = '.xlsx', count: int = 5):
    """Использует разметку и таблицу для создания новых картинок"""
    path = os.path.join(BASE_PATH, f"user_folders/{user_name}/{project_name}_table{table_format}")
    image_save_path = os.path.join(BASE_PATH, f"user_folders/{user_name}/{project_name}")

    try:
        workbook: Workbook = openpyxl.load_workbook(filename=path)
    except FileNotFoundError:
        return HTTPException(status_code=404, detail="table not found")

    # Получаем активный лист
    sheet: Worksheet = workbook.active

    indexes_for_printing = find_key_index(sheet, elements)
    index = 0
    element: ElementWrapper
    for row in sheet.iter_rows(values_only=True):
        if index > count: break

        try:
            image = Image.open(os.path.join(BASE_PATH, f"user_folders/{user_name}/{project_name}_image.jpg"))
        except FileNotFoundError:
            return HTTPException(status_code=404, detail="image not found")

        draw = ImageDraw.Draw(image)
        for element in elements.elements:
            font = ImageFont.truetype(find_ttf_file(element.element.font_family), 50)
            text = row[indexes_for_printing.get(element.key)]
            draw.text((element.position_x, element.position_y), text, font=font)
        # TODO: добавить еще стилизации
        image.save(image_save_path + f"{index}.png")
        index += 1

    # TODO: сделать нормальную выдачу фалов раз не получается архивом
    # await get_result(user_name, project_name)


if __name__ == '__main__':

    element = Element.parse_obj({})

    element_wrapper = ElementWrapper.parse_obj(
        {
            "key": "Тип операции",
            "element": element,
        }
    )
    some_elements = ElementsList.parse_obj({})

    some_elements.elements.append(
        element_wrapper
    )

    print(asyncio.run(print_excel_rows("a", "b", some_elements)))

